import openpyxl
import os
import fcntl
import html
import streamlit as st
from datetime import datetime

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "assets", "Digital_Currency.xlsx")

# ── Safe HTML helper — escapes all user-supplied strings before injecting ──────
def esc(value):
    """Escape a value for safe injection into HTML strings."""
    return html.escape(str(value) if value is not None else "")

COMMON_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800;900&family=Space+Mono:wght@400;700&display=swap');

/* ── Reset & base ── */
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}
html,body,[data-testid="stAppViewContainer"],[data-testid="stMain"]{
  background:#0f1117 !important;
  font-family:'Nunito',sans-serif;
  color:#e2e8f0;
}
[data-testid="stAppViewContainer"]{
  background:radial-gradient(ellipse 80% 40% at 50% -5%,rgba(99,102,241,0.18) 0%,transparent 65%),#0f1117 !important;
  min-height:100vh;
}
[data-testid="stHeader"]{background:transparent !important;box-shadow:none !important;}
#MainMenu,footer,[data-testid="stToolbar"]{visibility:hidden;}
[data-testid="stMainBlockContainer"]{
  max-width:500px !important;
  padding:0.75rem 0.75rem 6rem 0.75rem !important;
  margin:0 auto !important;
}

/* ── Tabs ── */
[data-testid="stTabs"] [role="tablist"]{
  background:#1e2130 !important;
  border-radius:14px !important;padding:5px !important;gap:4px !important;border:none !important;
}
[data-testid="stTabs"] [role="tab"]{
  font-family:'Nunito',sans-serif !important;font-weight:800 !important;font-size:0.88rem !important;
  color:#64748b !important;border-radius:10px !important;padding:10px 0 !important;
  flex:1 !important;text-align:center !important;transition:all 0.2s !important;border:none !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"]{
  background:#2a2f45 !important;color:#818cf8 !important;
  box-shadow:0 2px 8px rgba(0,0,0,0.4) !important;
}
[data-testid="stTabs"] [data-testid="stTabContent"]{padding-top:1rem !important;}

/* ── Inputs ── */
[data-testid="stTextInput"] label{
  font-family:'Nunito',sans-serif !important;font-size:0.75rem !important;
  font-weight:800 !important;letter-spacing:0.07em !important;
  text-transform:uppercase !important;color:#94a3b8 !important;
}
[data-testid="stTextInput"] input{
  background:#1e2130 !important;border:2px solid #2d3452 !important;
  border-radius:12px !important;color:#e2e8f0 !important;
  font-family:'Nunito',sans-serif !important;font-size:1rem !important;
  font-weight:700 !important;padding:12px 16px !important;
  height:50px !important;-webkit-appearance:none !important;
}
[data-testid="stTextInput"] input:focus{
  border-color:#6366f1 !important;
  box-shadow:0 0 0 4px rgba(99,102,241,0.2) !important;
  background:#252b40 !important;outline:none !important;
}
[data-testid="stTextInput"] input::placeholder{color:#475569 !important;}

/* ── Number input ── */
[data-testid="stNumberInput"] input{
  background:#1e2130 !important;border:2px solid #2d3452 !important;
  border-radius:12px !important;color:#e2e8f0 !important;
}
[data-testid="stNumberInput"] button{
  background:#2a2f45 !important;border-color:#2d3452 !important;color:#94a3b8 !important;
}

/* ── Selectbox ── */
[data-testid="stSelectbox"] > div > div{
  background:#1e2130 !important;border:2px solid #2d3452 !important;
  border-radius:12px !important;color:#e2e8f0 !important;
}

/* ── Radio ── */
[data-testid="stRadio"] label{color:#94a3b8 !important;font-weight:700 !important;}
[data-testid="stRadio"] [data-testid="stMarkdownContainer"] p{color:#e2e8f0 !important;}

/* ── Expander ── */
[data-testid="stExpander"]{
  background:#1e2130 !important;border:1px solid #2d3452 !important;
  border-radius:14px !important;
}
[data-testid="stExpander"] summary{color:#94a3b8 !important;font-weight:700 !important;}

/* ── PIN dots ── */
.pin-dots{display:flex;justify-content:center;gap:14px;margin:0.5rem 0 0.75rem 0;}
.pin-dot{
  width:16px;height:16px;border-radius:50%;
  border:2.5px solid #2d3452;background:transparent;transition:all 0.18s ease;
}
.pin-dot.filled{
  background:#6366f1;border-color:#6366f1;
  box-shadow:0 2px 8px rgba(99,102,241,0.5);transform:scale(1.2);
}

/* ── Buttons ── */
[data-testid="stButton"] button{
  font-family:'Nunito',sans-serif !important;font-weight:800 !important;
  font-size:0.95rem !important;
  background:linear-gradient(135deg,#4f46e5 0%,#7c3aed 100%) !important;
  color:white !important;border:none !important;border-radius:14px !important;
  padding:0 !important;height:52px !important;width:100% !important;
  cursor:pointer !important;box-shadow:0 4px 20px rgba(79,70,229,0.4) !important;
  margin-top:0.5rem !important;-webkit-tap-highlight-color:transparent !important;
  transition:opacity 0.2s,transform 0.15s,box-shadow 0.2s !important;
}
[data-testid="stButton"] button:hover{
  box-shadow:0 6px 28px rgba(79,70,229,0.55) !important;transform:translateY(-1px) !important;
}
[data-testid="stButton"] button:active{transform:translateY(0) scale(0.98) !important;}

/* ── Alerts ── */
[data-testid="stSuccess"]{
  background:#052e16 !important;border:1.5px solid #166534 !important;
  border-radius:12px !important;color:#86efac !important;font-weight:700 !important;
}
[data-testid="stError"]{
  background:#2d0a0a !important;border:1.5px solid #7f1d1d !important;
  border-radius:12px !important;color:#fca5a5 !important;font-weight:700 !important;
}
[data-testid="stWarning"]{
  background:#1c1204 !important;border:1.5px solid #78350f !important;
  border-radius:12px !important;color:#fcd34d !important;font-weight:700 !important;
}
[data-testid="stInfo"]{
  background:#0c1a2e !important;border:1.5px solid #1e3a5f !important;
  border-radius:12px !important;color:#93c5fd !important;font-weight:700 !important;
}

/* ── Spinner ── */
[data-testid="stSpinner"]{color:#818cf8 !important;}

/* ── Custom components ── */
.top-bar{
  background:linear-gradient(135deg,#4f46e5,#7c3aed);
  border-radius:20px;padding:1rem 1.2rem;color:white;
  box-shadow:0 8px 32px rgba(79,70,229,0.35);margin-bottom:0.75rem;
}
.pin-dots{display:flex;justify-content:center;gap:14px;margin:0.5rem 0 0.75rem 0;}

.card{
  background:#1e2130;border-radius:16px;padding:1rem 1.1rem;
  margin-bottom:0.6rem;box-shadow:0 2px 12px rgba(0,0,0,0.3);
  border-left:4px solid #4f46e5;
}
.card-orange{border-left-color:#d97706;}
.card-green{border-left-color:#059669;}
.card-red{border-left-color:#dc2626;}
.card-yellow{border-left-color:#f59e0b;}
.card-center{text-align:center;}

.stall-card{
  background:#1e2130;border-radius:16px;padding:1rem 1.1rem;
  margin-bottom:0.6rem;box-shadow:0 2px 12px rgba(0,0,0,0.3);
  border:1px solid #2d3452;
}
.item-row{
  display:flex;justify-content:space-between;
  padding:6px 0;border-bottom:1px solid #1e2130;
}
.cart-bar{
  background:#1a1f35;border-radius:16px;padding:1rem 1.1rem;
  margin:0.75rem 0;border:1.5px solid #2d3452;
}
.cart-title{
  font-size:0.75rem;font-weight:800;text-transform:uppercase;
  letter-spacing:0.08em;color:#818cf8;margin-bottom:0.5rem;
}
.cart-row{
  display:flex;justify-content:space-between;
  font-size:0.88rem;font-weight:700;color:#cbd5e1;padding:3px 0;
}
.cart-total{
  border-top:1.5px solid #2d3452;margin-top:0.5rem;padding-top:0.5rem;
  font-size:1rem;font-weight:900;color:#e2e8f0;
  display:flex;justify-content:space-between;
}
.token-card{
  background:linear-gradient(135deg,#065f46,#047857);
  border-radius:24px;padding:2rem 1.5rem;color:white;text-align:center;
  box-shadow:0 8px 32px rgba(5,150,105,0.4);
}
.token-num{
  font-family:'Space Mono',monospace;font-size:3rem;
  font-weight:700;letter-spacing:0.1em;
}
.info-row{
  display:flex;justify-content:space-between;align-items:center;
  background:#1e2130;border-radius:12px;padding:0.85rem 1rem;
  margin-bottom:0.5rem;border:1px solid #2d3452;
}
.info-label{
  font-size:0.7rem;font-weight:800;text-transform:uppercase;
  letter-spacing:0.07em;color:#64748b;
}
.info-value{
  font-size:0.85rem;font-weight:800;color:#e2e8f0;
  font-family:'Space Mono',monospace;
}
.hint{
  font-size:0.73rem;color:#475569;font-weight:600;
  margin-top:0.25rem;text-align:center;
}
.section-title{
  font-size:0.95rem;font-weight:900;color:#cbd5e1;
  margin:0.75rem 0 0.5rem 0;letter-spacing:-0.01em;
}
.green-badge{
  display:inline-block;background:#052e16;color:#86efac;
  font-size:0.68rem;font-weight:800;padding:3px 10px;border-radius:20px;
  border:1px solid #166534;
}
.red-badge{
  display:inline-block;background:#2d0a0a;color:#fca5a5;
  font-size:0.68rem;font-weight:800;padding:3px 10px;border-radius:20px;
  border:1px solid #7f1d1d;
}
.orange-badge{
  display:inline-block;background:#1c0f00;color:#fed7aa;
  font-size:0.68rem;font-weight:800;padding:3px 10px;border-radius:20px;
  border:1px solid #78350f;
}
.blue-badge{
  display:inline-block;background:#0c1a2e;color:#93c5fd;
  font-size:0.68rem;font-weight:800;padding:3px 10px;border-radius:20px;
  border:1px solid #1e3a5f;
}
</style>
"""

# ══════════════════════════════════════════════════════════════════════════════
# FILE LOCKING — fcntl exclusive lock for all writes
# ══════════════════════════════════════════════════════════════════════════════
class ExcelLock:
    def __init__(self):
        self.lock_path = EXCEL_PATH + ".lock"
        self.fh = None
    def __enter__(self):
        self.fh = open(self.lock_path, "w")
        fcntl.flock(self.fh, fcntl.LOCK_EX)
        return self
    def __exit__(self, *_):
        fcntl.flock(self.fh, fcntl.LOCK_UN)
        self.fh.close()

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS — always read fresh (no caching at all)
# ══════════════════════════════════════════════════════════════════════════════
def load_wb():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=True)

def load_wb_readonly():
    """Faster read-only load for queries that don't write."""
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

def sheet_to_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row): continue
        result.append(dict(zip(headers, row)))
    return result

def ensure_tokens_sheet(wb):
    if "Tokens" not in wb.sheetnames:
        ws = wb.create_sheet("Tokens")
        ws.append(["Token_No","Stall_ID","Stall_Name","Username","Items","Total","Time","Status"])
    return wb

# ══════════════════════════════════════════════════════════════════════════════
# USER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def get_all_users():
    wb = load_wb_readonly()
    data = sheet_to_dicts(wb["Users"])
    wb.close()
    return data

def find_user(username):
    for u in get_all_users():
        if str(u.get("Username","")).lower() == username.lower():
            return u
    return None

def get_live_balance(username):
    u = find_user(username)
    try: return float(u.get("Amount", 0) or 0) if u else 0
    except: return 0

def get_next_uid():
    max_num = 0
    for u in get_all_users():
        uid = str(u.get("UID",""))
        if uid.startswith("UID_"):
            try:
                n = int(uid.split("_")[1])
                if n > max_num: max_num = n
            except: pass
    return f"UID_{str(max_num + 1).zfill(4)}"

def register_user(username, pin):
    with ExcelLock():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        # Re-check username uniqueness inside the lock
        for row in wb["Users"].iter_rows(min_row=2, values_only=True):
            if row[1] and str(row[1]).lower() == username.lower():
                wb.close()
                return None, "Username already taken."
        uid = "UID_0001"
        max_num = 0
        for row in wb["Users"].iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).startswith("UID_"):
                try:
                    n = int(str(row[0]).split("_")[1])
                    if n > max_num: max_num = n
                except: pass
        uid = f"UID_{str(max_num + 1).zfill(4)}"
        wb["Users"].append([uid, username, int(pin), 0])
        wb.save(EXCEL_PATH)
        wb.close()
    return {"UID": uid, "Username": username, "Pin": pin, "Amount": 0}, None

def update_balance(username, new_amount):
    with ExcelLock():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Users"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        amt_col = headers.index("Amount") + 1
        usr_col = headers.index("Username") + 1
        for i in range(2, ws.max_row + 1):
            if str(ws.cell(i, usr_col).value or "").lower() == username.lower():
                ws.cell(row=i, column=amt_col, value=round(float(new_amount), 2))
                wb.save(EXCEL_PATH)
                wb.close()
                return True
        wb.close()
    return False

# ══════════════════════════════════════════════════════════════════════════════
# TOKEN + PAYMENT  — single atomic locked operation
# ══════════════════════════════════════════════════════════════════════════════
def save_token_and_deduct(username, stall_id, stall_name, items_dict, total):
    """
    Single lock: verify balance → deduct → assign token number → write row.
    Returns (token_no, new_balance) or (None, current_balance) on failure.
    """
    with ExcelLock():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        wb = ensure_tokens_sheet(wb)

        # 1. Find user and check live balance
        ws_u = wb["Users"]
        u_headers = [c.value for c in next(ws_u.iter_rows(min_row=1, max_row=1))]
        amt_col = u_headers.index("Amount") + 1
        usr_col = u_headers.index("Username") + 1
        new_bal = None
        for i in range(2, ws_u.max_row + 1):
            cell_user = str(ws_u.cell(i, usr_col).value or "")
            if cell_user.lower() == username.lower():
                cur = float(ws_u.cell(i, amt_col).value or 0)
                if cur < total:
                    wb.close()
                    return None, cur
                new_bal = round(cur - total, 2)
                ws_u.cell(row=i, column=amt_col, value=new_bal)
                break

        if new_bal is None:
            wb.close()
            return None, 0

        # 2. Get next token number
        ws_t = wb["Tokens"]
        max_tok = 0
        for row in ws_t.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    n = int(row[0])
                    if n > max_tok: max_tok = n
                except: pass
        token_no = max_tok + 1

        # 3. Write token row
        items_str = ", ".join([f"{v['qty']}x {k} (₹{int(v['price']*v['qty'])})" for k, v in items_dict.items()])
        timestamp = datetime.now().strftime("%d %b %Y, %I:%M %p")
        ws_t.append([token_no, stall_id, stall_name, username, items_str, round(total, 2), timestamp, "Pending"])

        wb.save(EXCEL_PATH)
        wb.close()
        return token_no, new_bal

def get_tokens_for_user(username):
    wb = load_wb_readonly()
    if "Tokens" not in wb.sheetnames:
        wb.close(); return []
    data = [t for t in sheet_to_dicts(wb["Tokens"])
            if str(t.get("Username","")).lower() == username.lower()]
    wb.close()
    return data

def get_tokens_for_stall(stall_id):
    wb = load_wb_readonly()
    if "Tokens" not in wb.sheetnames:
        wb.close(); return []
    data = [t for t in sheet_to_dicts(wb["Tokens"])
            if str(t.get("Stall_ID","")).strip() == str(stall_id).strip()]
    wb.close()
    return data

def get_all_tokens():
    wb = load_wb_readonly()
    if "Tokens" not in wb.sheetnames:
        wb.close(); return []
    data = sheet_to_dicts(wb["Tokens"])
    wb.close()
    return data

def mark_token_served(token_no, stall_id):
    with ExcelLock():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Tokens" not in wb.sheetnames:
            wb.close(); return None
        ws = wb["Tokens"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        tc  = headers.index("Token_No") + 1
        sc  = headers.index("Stall_ID") + 1
        stc = headers.index("Status") + 1
        for i in range(2, ws.max_row + 1):
            if (str(ws.cell(i, tc).value) == str(token_no) and
                    str(ws.cell(i, sc).value or "").strip() == str(stall_id).strip()):
                cur   = str(ws.cell(i, stc).value or "Pending")
                new_s = "Served" if cur != "Served" else "Pending"
                ws.cell(row=i, column=stc, value=new_s)
                wb.save(EXCEL_PATH)
                wb.close()
                return new_s
        wb.close()
    return None

# ══════════════════════════════════════════════════════════════════════════════
# STALL FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def load_stalls():
    wb = load_wb_readonly()
    stalls = {}
    for row in wb["Stalls"].iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        sid  = str(row[0]).strip()
        name = str(row[1]).strip()
        pin  = str(row[2]).strip()
        stalls[sid] = {"id": sid, "name": name, "pin": pin, "menu": []}
    for row in wb["Menu"].iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        sid  = str(row[0]).strip()
        item = str(row[2]).strip()
        try: price = float(row[3])
        except: price = 0
        if sid in stalls:
            stalls[sid]["menu"].append({"item": item, "price": price})
    wb.close()
    return list(stalls.values())

def find_stall(stall_id, stall_name, pin):
    for s in load_stalls():
        if (s["id"].lower()   == stall_id.strip().lower() and
                s["name"].lower() == stall_name.strip().lower() and
                str(s["pin"]).strip() == str(pin).strip()):
            return s
    return None

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def get_admin_credentials():
    wb = load_wb_readonly()
    rows = list(wb["Counter_Admin"].iter_rows(min_row=2, values_only=True))
    wb.close()
    if rows and rows[0][0]:
        return str(rows[0][0]).strip(), str(rows[0][1]).strip()
    return None, None

def block_user(username):
    with ExcelLock():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Users"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        pin_col = headers.index("Pin") + 1
        usr_col = headers.index("Username") + 1
        for i in range(2, ws.max_row + 1):
            if str(ws.cell(i, usr_col).value or "").lower() == username.lower():
                ws.cell(row=i, column=pin_col, value="BLOCKED")
                wb.save(EXCEL_PATH)
                wb.close()
                return True
        wb.close()
    return False

def unblock_user(username, new_pin):
    with ExcelLock():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Users"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        pin_col = headers.index("Pin") + 1
        usr_col = headers.index("Username") + 1
        for i in range(2, ws.max_row + 1):
            if str(ws.cell(i, usr_col).value or "").lower() == username.lower():
                ws.cell(row=i, column=pin_col, value=int(new_pin))
                wb.save(EXCEL_PATH)
                wb.close()
                return True
        wb.close()
    return False

# ══════════════════════════════════════════════════════════════════════════════
# SHARED UI HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def render_pin_dots(filled):
    dots = "".join(
        f'<div class="pin-dot{"  filled" if i < filled else ""}"></div>'
        for i in range(4)
    )
    st.markdown(f'<div class="pin-dots">{dots}</div>', unsafe_allow_html=True)

def top_bar(label, name, sub, balance=None, color="#4f46e5"):
    try:    bal_str = f"₹{float(balance):,.2f}"
    except: bal_str = str(balance) if balance is not None else ""
    bal_html = (
        f'<div style="text-align:right;">'
        f'<div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;opacity:0.7;">Balance</div>'
        f'<div style="font-family:\'Space Mono\',monospace;font-size:1.4rem;font-weight:700;">{esc(bal_str)}</div>'
        f'</div>'
    ) if balance is not None else ""

    st.markdown(f"""
    <div class="top-bar" style="background:linear-gradient(135deg,{esc(color)},{esc(color)}cc);">
      <div style="display:flex;justify-content:space-between;align-items:center;">
        <div>
          <div style="font-size:0.65rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;opacity:0.7;">{esc(label)}</div>
          <div style="font-size:1.15rem;font-weight:900;">{esc(name)}</div>
          <div style="font-size:0.6rem;opacity:0.5;font-family:'Space Mono',monospace;">{esc(sub)}</div>
        </div>
        {bal_html}
      </div>
    </div>
    """, unsafe_allow_html=True)

def stat_card(value, label, color="#818cf8", extra_style=""):
    st.markdown(f"""
    <div class="card card-center" style="border-left-color:{esc(color)};{extra_style}">
      <div style="font-size:1.6rem;font-weight:900;color:{esc(color)};">{esc(str(value))}</div>
      <div style="font-size:0.65rem;font-weight:800;text-transform:uppercase;color:#64748b;margin-top:2px;">{esc(label)}</div>
    </div>""", unsafe_allow_html=True)
